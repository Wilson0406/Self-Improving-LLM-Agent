# import streamlit as st
# import pandas as pd
# from PyPDF2 import PdfReader
# from openpyxl import load_workbook
# from extraction_agent import call_extraction_agent
# import asyncio

# st.set_page_config(page_title="Document Extraction Chatbot", layout="wide")
# st.sidebar.title("Upload & Options")

# pdf_file = st.sidebar.file_uploader("Upload PDF", type=["pdf"])
# excel_file = st.sidebar.file_uploader("Upload Excel (column schema)", type=["xlsx"])
# show_feedback_log = st.sidebar.checkbox("Show Feedback Log", value=True)

# if "feedback_log" not in st.session_state:
#     st.session_state.feedback_log = []
# if "chat_history" not in st.session_state:
#     st.session_state.chat_history = []

# st.title("Document Extraction Chatbot")

# def extract_text_from_pdf(pdf_file):
#     reader = PdfReader(pdf_file)
#     return "\n".join([page.extract_text() or "" for page in reader.pages])

# def read_excel_schema(excel_file):
#     wb = load_workbook(excel_file)
#     ws = wb.active
#     columns = [cell.value for cell in ws[1]]
#     instructions = [cell.value for cell in ws[2]]
#     return columns, instructions

# # Display PDF/Excel previews
# if pdf_file:
#     st.markdown("**PDF Uploaded!**")
#     pdf_text = extract_text_from_pdf(pdf_file)
#     st.text_area("PDF Preview", value=pdf_text[:200], height=100)
# else:
#     pdf_text = None

# if excel_file:
#     st.markdown("**Excel Uploaded!**")
#     columns, instructions = read_excel_schema(excel_file)
#     df = pd.DataFrame([instructions], columns=columns)
#     st.write(df)
# else:
#     columns, instructions = None, None

# # Main Chat UI
# with st.container():
#     st.markdown("### Chat Interface")
#     for entry in st.session_state.chat_history:
#         st.write(f"**You:** {entry['user']}")
#         st.write(f"**Bot:** {entry['bot']}")

#     user_input = st.text_input(
#         "Ask questions, give extraction directives, or clarify schema/instructions...", key="chat_input"
#     )

#     if st.button("Send", key="send_button"):
#         # Validate files
#         if not pdf_text or not columns or not instructions:
#             st.error("Upload both PDF and Excel files (with columns/instructions) before chatting!")
#         else:
#             # Call backend agent with current context
#             bot_response = asyncio.run(
#             call_extraction_agent(pdf_text, columns, instructions)
#         )
#             st.session_state.chat_history.append({"user": user_input, "bot": bot_response})
#             st.rerun()

#     # Add feedback for each response
#     feedback = st.text_area("Suggest correction/feedback for last extraction?", key="feedback_box")
#     if st.button("Submit Feedback", key="feedback_button"):
#         if st.session_state.chat_history:
#             query = st.session_state.chat_history[-1]["user"]
#             st.session_state.feedback_log.append({"query": query, "feedback": feedback})
#             st.success("Feedback logged!")

# # Feedback log in sidebar
# if show_feedback_log:
#     st.sidebar.markdown("### Feedback Log")
#     st.sidebar.write(st.session_state.feedback_log)

import streamlit as st
import pandas as pd
from PyPDF2 import PdfReader
from openpyxl import load_workbook
import asyncio
from extraction_agent import call_extraction_agent

st.set_page_config(page_title="Document Extraction Chatbot", layout="wide")
st.sidebar.title("Upload Section (Mandatory)")

# --- UPLOAD SECTION: SINGLE GROUPED FORM ---
with st.sidebar.form("file_upload_form"):
    pdf_file = st.file_uploader("Upload PDF", type=["pdf"], key="pdf_file")
    excel_file = st.file_uploader("Upload Excel (column schema)", type=["xlsx"], key="excel_file")
    file_submit = st.form_submit_button("Confirm Uploads")

# --- CHECK IF BOTH FILES PRESENT ---
if not pdf_file or not excel_file:
    st.warning("Please upload **both** a PDF **and** an Excel file to proceed.")
    files_uploaded = False
else:
    files_uploaded = True

# --- SESSION STATE INIT ---
if "feedback_log" not in st.session_state:
    st.session_state.feedback_log = []
if "chat_history" not in st.session_state:
    st.session_state.chat_history = []

st.title("Document Extraction Chatbot")

# --- FILE UTILITIES ---
def extract_text_from_pdf(pdf_file):
    reader = PdfReader(pdf_file)
    return "\n".join([page.extract_text() or "" for page in reader.pages])

def read_excel_schema(excel_file):
    wb = load_workbook(excel_file)
    ws = wb.active
    columns = [cell.value for cell in ws[1]]
    instructions = [cell.value for cell in ws[2]]
    return columns, instructions

# --- MAIN SECTION: ONLY RUN IF FILES UPLOADED ---
if files_uploaded:
    st.markdown("**PDF Uploaded!**")
    pdf_text = extract_text_from_pdf(pdf_file)
    st.text_area("PDF Preview", value=pdf_text[:300], height=100)

    st.markdown("**Excel Uploaded!**")
    columns, instructions = read_excel_schema(excel_file)
    df = pd.DataFrame([instructions], columns=columns)
    st.write(df)

    # --- CHAT BLOCK ---
    with st.container():
        st.markdown("### Chat Interface")
        for entry in st.session_state.chat_history:
            st.write(f"**You:** {entry['user']}")
            st.write(f"**Bot:** {entry['bot']}")

        user_input = st.text_input(
            "Ask extraction question or give directives...", key="chat_input"
        )

        if st.button("Send", key="send_button") and user_input:
            # Call backend agent (ADK/GPT-4o)
            bot_response = asyncio.run(
                call_extraction_agent(pdf_text, columns, instructions)
            )
            st.session_state.chat_history.append({"user": user_input, "bot": bot_response})
            st.rerun()

        # Feedback for last
        feedback = st.text_area("Suggest correction/feedback for last extraction?", key="feedback_box")
        if st.button("Submit Feedback", key="feedback_button"):
            if st.session_state.chat_history:
                query = st.session_state.chat_history[-1]["user"]
                st.session_state.feedback_log.append({"query": query, "feedback": feedback})
                st.success("Feedback logged!")

else:
    st.info("Please finish uploading your PDF and Excel file before starting extraction or chat.")

# --- FEEDBACK LOG ---
if st.sidebar.checkbox("Show Feedback Log", value=True):
    st.sidebar.markdown("### Feedback Log")
    st.sidebar.write(st.session_state.feedback_log)
