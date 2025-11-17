import streamlit as st
import pandas as pd
from PyPDF2 import PdfReader
from openpyxl import load_workbook
import asyncio
from extraction_agent import call_extraction_agent
from improvement_agent import call_improvement_agent
from document_intelligence import doc_intelligence
import json
import io
import os
import getpass
from datetime import datetime
from database import DatabaseManager, get_latest_prompt, save_improved_prompt

st.set_page_config(page_title="Document Extraction Feedback", layout="wide")
st.sidebar.title("Upload Section (Mandatory)")

with st.sidebar.form("file_upload_form"):
    use_case = st.selectbox(
        "Select Use Case",
        options=["Form 926", "Form 1040"],
        index=0,
        help="Choose the type of document you're processing"
    )
    pdf_file = st.file_uploader("Upload PDF", type=["pdf"], key="pdf_file")
    excel_file = st.file_uploader("Upload Excel (column schema)", type=["xlsx"], key="excel_file")
    file_submit = st.form_submit_button("Confirm Uploads")

if not pdf_file or not excel_file:
    st.warning("Please upload **both** a PDF **and** an Excel file to proceed.")
    files_uploaded = False
else:
    files_uploaded = True
    # Store the selected use case in session state
    st.session_state['use_case'] = use_case

if "feedback_log" not in st.session_state:
    st.session_state.feedback_log = []
if "last_extraction" not in st.session_state:
    st.session_state['last_extraction'] = None
if "last_prompt" not in st.session_state:
    st.session_state['last_prompt'] = None
if "improved_prompt" not in st.session_state:
    st.session_state['improved_prompt'] = None
if "use_case" not in st.session_state:
    st.session_state['use_case'] = "Form 926"
if "improved_extraction" not in st.session_state:
    st.session_state['improved_extraction'] = None
if "db_manager" not in st.session_state:
    st.session_state['db_manager'] = None
if "current_prompt_id" not in st.session_state:
    st.session_state['current_prompt_id'] = None
if "document_id" not in st.session_state:
    st.session_state['document_id'] = None
if "document_versions" not in st.session_state:
    st.session_state['document_versions'] = []
if "current_document_id" not in st.session_state:
    st.session_state['current_document_id'] = None

# Detect and store a local user id for local runs (used as user_id in DB operations)
if 'user_id' not in st.session_state:
    detected_user = None
    for key in ("USER", "USERNAME", "LOGNAME"):
        val = os.environ.get(key)
        if val:
            detected_user = val
            break
    if not detected_user:
        try:
            detected_user = getpass.getuser()
        except Exception:
            detected_user = None
    if not detected_user:
        try:
            detected_user = os.getlogin()
        except Exception:
            detected_user = None
    st.session_state['user_id'] = detected_user or 'local_user'

st.title("Document Extraction Feedback + Prompt Refinement")

def extract_text_from_pdf(pdf_file):
    # Get file content as bytes and pass directly to doc_intelligence function
    file_content = pdf_file.getvalue()

    # Use doc_intelligence function to extract text with Azure Document Intelligence
    extracted_content = doc_intelligence(file_content)
    
    return extracted_content

def read_excel_schema(excel_file):
    wb = load_workbook(excel_file)
    ws = wb.active
    columns = [cell.value for cell in ws[1]]
    
    # Check if second row exists and has data
    instructions = []
    if ws.max_row >= 2:
        # Check if second row has any non-empty cells
        second_row_values = [cell.value for cell in ws[2]]
        if any(value is not None and str(value).strip() for value in second_row_values):
            instructions = second_row_values
        else:
            # If second row is empty, use empty strings as placeholders
            instructions = ["" for _ in columns]
    else:
        # If no second row exists, use empty strings as placeholders
        instructions = ["" for _ in columns]
    
    return columns, instructions

def json_to_excel(json_data, columns):
    """Convert JSON data to Excel format and return as bytes"""
    try:
        # Parse JSON if it's a string
        if isinstance(json_data, str):
            data = json.loads(json_data)
        else:
            data = json_data
        
        # Create DataFrame with the extracted data
        df = pd.DataFrame([data])
        
        # Ensure all expected columns are present, fill missing ones with empty strings
        for col in columns:
            if col not in df.columns:
                df[col] = ""
        
        # Reorder columns to match the original schema
        df = df[columns]
        
        # Create Excel file in memory
        output = io.BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name='Extracted_Data')
        
        return output.getvalue()
    
    except json.JSONDecodeError:
        # If JSON parsing fails, create a simple error sheet
        error_df = pd.DataFrame({"Error": ["Invalid JSON format in extracted data"]})
        output = io.BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            error_df.to_excel(writer, index=False, sheet_name='Error')
        return output.getvalue()

def json_to_dataframe(json_data, columns):
    """Convert JSON extraction data to DataFrame for display"""
    try:
        # Parse JSON if it's a string
        if isinstance(json_data, str):
            data = json.loads(json_data)
        else:
            data = json_data
        
        # Create DataFrame with the extracted data
        df = pd.DataFrame([data])
        
        # Ensure all expected columns are present, fill missing ones with empty strings
        for col in columns:
            if col not in df.columns:
                df[col] = ""
        
        # Reorder columns to match the original schema
        df = df[columns]
        
        return df
    
    except json.JSONDecodeError:
        # If JSON parsing fails, create an error DataFrame
        error_df = pd.DataFrame({"Error": ["Invalid JSON format in extracted data"]})
        return error_df
    except Exception as e:
        # Handle any other errors
        error_df = pd.DataFrame({"Error": [f"Error processing data: {str(e)}"]})
        return error_df

if files_uploaded:
    # st.markdown("**PDF Preview**")
    pdf_text = extract_text_from_pdf(pdf_file)
    # st.text_area("PDF Preview", value=pdf_text[:500], height=120)

    st.markdown("**Excel Schema & Instructions**")
    columns, instructions = read_excel_schema(excel_file)
    df = pd.DataFrame([instructions], columns=columns)
    st.write(df)

    # --- Initialize Database Manager ---
    if st.session_state['db_manager'] is None:
        try:
            st.session_state['db_manager'] = DatabaseManager()
            # Test connection
            if st.session_state['db_manager'].test_connection():
                st.success("✅ Database connected successfully")
            else:
                st.warning("⚠️ Database connection failed - using local prompts")
                st.session_state['db_manager'] = None
        except Exception as e:
            st.warning(f"⚠️ Database initialization failed: {str(e)} - using local prompts")
            st.session_state['db_manager'] = None

    # --- Manage Document in Database ---
    if st.session_state['db_manager'] and st.session_state.get("last_uploaded_files") != (pdf_file.name, excel_file.name):
        try:
            # Always create a new document record for each submit (preserves full history)
            # Check existing documents only for information/reference
            existing_docs = st.session_state['db_manager'].get_document_by_filename(pdf_file.name)
            
            if existing_docs:
                st.info(f"📄 Previous submissions found for '{pdf_file.name}'. Creating new record for this session.")
                
                # Show existing document info for reference
                with st.expander(f"📋 Previous Submissions for '{pdf_file.name}'"):
                    st.write(f"**Last Document ID:** {existing_docs['DocumentID']}")
                    st.write(f"**Last Status:** {existing_docs.get('ExtractionStatus', 'N/A')}")
                    st.write(f"**Last Updated:** {existing_docs.get('LastUpdated', 'N/A')}")
                    st.caption("A new record will be created for this session to preserve complete history.")
            
            # Always insert new document request for each submit
            with st.spinner("📄 Creating new document record in database..."):
                try:
                    # determine file extension as source_type
                    _, ext = os.path.splitext(pdf_file.name)
                    source_type = ext.lstrip('.').lower() if ext else 'pdf'
                    
                    # Add timestamp to source_type to make each submission unique
                    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                    # source_type = f"{source_type}_submit_{timestamp}"

                    # Use existing user_id from session state (already detected during initialization)
                    local_user = st.session_state.get('user_id', 'streamlit_user')

                    document_id = st.session_state['db_manager'].insert_document_request(
                        file_name=pdf_file.name,
                        user_id=local_user,
                        source_type=source_type
                    )
                    if document_id:
                        st.session_state['document_id'] = document_id
                        
                        # Reset document versions for new submission
                        st.session_state['document_versions'] = []
                        st.session_state['current_document_id'] = document_id
                        
                        st.success(f"📄 New submission record created for '{pdf_file.name}' (ID: {document_id})")
                        st.info(f"🔖 Source Type: {source_type}")
                    else:
                        st.error("❌ Failed to register document in database - stored procedure returned NULL")
                        st.session_state['document_id'] = None
                except Exception as insert_e:
                    st.error(f"❌ Document insertion error: {str(insert_e)}")
                    st.session_state['document_id'] = None
        except Exception as e:
            st.error(f"❌ Document management failed: {str(e)}")
            st.session_state['document_id'] = None

    # --- Get prompt from database or build locally ---
    if st.session_state['db_manager']:
        try:
            # Try to get prompt from database
            db_prompt_data = get_latest_prompt(st.session_state.get('use_case', 'Form 926'))
            if db_prompt_data:
                # Use database prompt as base and customize with current data
                base_prompt = db_prompt_data['PromptText']
                st.session_state['current_prompt_id'] = db_prompt_data['PromptID']
                
                # Customize the prompt with current extraction details
                prompt = f"{base_prompt}\n\n" + (
                    f"Current Task:\n"
                    f"Extract the following columns from the PDF text:\n"
                    f"Columns: {', '.join(columns)}\n"
                    f"Instructions: {'; '.join(instructions)}\n"
                    f"PDF Content:\n{pdf_text}\n"
                )
                
                st.info(f"📋 Using database prompt (ID: {db_prompt_data['PromptID']}) - {db_prompt_data['PromptTitle']}")
            else:
                # Fallback to local prompt
                prompt = (
                    f"Extract the following columns from the PDF text:\n"
                    f"Columns: {', '.join(columns)}\n"
                    f"Instructions: {'; '.join(instructions)}\n"
                    f"PDF Content:\n{pdf_text}\n"
                    "Return your answer as a JSON object. "
                    "Do not write any commentary—output only the JSON in this format: "
                    '{"column1": "extractedValue", ...}'
                )
                st.info("📝 No database prompt found - using default prompt")
        except Exception as e:
            st.warning(f"⚠️ Database prompt fetch failed: {str(e)} - using local prompt")
            # Fallback to local prompt
            prompt = (
                f"Extract the following columns from the PDF text:\n"
                f"Columns: {', '.join(columns)}\n"
                f"Instructions: {'; '.join(instructions)}\n"
                f"PDF Content:\n{pdf_text}\n"
                "Return your answer as a JSON object. "
                "Do not write any commentary—output only the JSON in this format: "
                '{"column1": "extractedValue", ...}'
            )
    else:
        # Build local prompt when database is not available
        prompt = (
            f"Extract the following columns from the PDF text:\n"
            f"Columns: {', '.join(columns)}\n"
            f"Instructions: {'; '.join(instructions)}\n"
            f"PDF Content:\n{pdf_text}\n"
            "Return your answer as a JSON object. "
            "Do not write any commentary—output only the JSON in this format: "
            '{"column1": "extractedValue", ...}'
        )
    
    st.session_state['last_prompt'] = prompt

    # --- Run extraction if new files uploaded OR if no extraction exists yet ---
    need_extraction = (
        st.session_state.get("last_uploaded_files") != (pdf_file.name, excel_file.name) or
        st.session_state.get('last_extraction') is None
    )
    
    if need_extraction:
        with st.spinner("🔄 Running extraction..."):
            st.session_state['last_extraction'] = asyncio.run(
                call_extraction_agent(prompt, columns, instructions)
            )
        
        st.session_state['improved_prompt'] = None
        st.session_state['improved_extraction'] = None
        st.session_state["last_uploaded_files"] = (pdf_file.name, excel_file.name)
    
    # --- Always save extraction results to database if we have document_id and extraction ---
    if (st.session_state['db_manager'] and 
        st.session_state.get('document_id') and 
        st.session_state.get('last_extraction') and
        need_extraction):
        
        try:
            with st.spinner("💾 Saving extraction results to database..."):
                success = st.session_state['db_manager'].update_document_master_by_id(
                    document_id=st.session_state['document_id'],
                    extraction_status="Completed",
                    extraction_output=st.session_state['last_extraction'],
                    prompt_id=st.session_state.get('current_prompt_id'),
                    retry_count=0,
                    error_message=None,
                    comments=f"Initial extraction completed via Streamlit interface. File: {pdf_file.name}, Columns: {len(columns)}"
                )
                if success:
                    st.success(f"✅ Extraction results saved to database (Document ID: {st.session_state['document_id']})")
                    
                    # Initialize document version tracking
                    if 'document_versions' not in st.session_state:
                        st.session_state['document_versions'] = []
                    
                    # Track this as version 1 (initial extraction)
                    version_info = {
                        'version': 1,
                        'document_id': st.session_state['document_id'],
                        'extraction_type': 'Initial',
                        'feedback': None
                    }
                    st.session_state['document_versions'] = [version_info]  # Reset for new document
                    
                    # Show what was saved for verification
                    with st.expander("📋 Database Save Details"):
                        st.write(f"**Document ID:** {st.session_state['document_id']}")
                        st.write(f"**Version:** 1 (Initial)")
                        st.write(f"**Status:** Completed")
                        st.write(f"**Prompt ID:** {st.session_state.get('current_prompt_id', 'N/A')}")
                        st.write(f"**Extraction Output:** {len(st.session_state['last_extraction'])} characters")
                        st.write(f"**Comments:** Initial extraction for {pdf_file.name}")
                else:
                    st.error("❌ Failed to save extraction results to database")
        except Exception as e:
            st.error(f"❌ Database update failed: {str(e)}")
            # Update status to error in database
            try:
                st.session_state['db_manager'].update_document_master_by_id(
                    document_id=st.session_state['document_id'],
                    extraction_status="Error",
                    error_message=f"Initial extraction error: {str(e)}",
                    retry_count=0,
                    comments=f"Error during initial extraction for {pdf_file.name}"
                )
                st.warning("⚠️ Document status updated to 'Error' in database")
            except Exception as nested_e:
                st.error(f"❌ Could not update error status: {str(nested_e)}")

    # --- Show extraction ---
    st.markdown("### Extraction Agent Output (Initial)")
    
    # Display as table
    extraction_df = json_to_dataframe(st.session_state['last_extraction'], columns)
    st.dataframe(extraction_df, use_container_width=True)
    
    # Option to view raw JSON
    with st.expander("🔍 View Raw JSON Output"):
        st.code(st.session_state['last_extraction'], language="json")

    # --- Feedback section + RUN IMPROVEMENT AGENT ---
    st.markdown("### Feedback or Correction")
    
    # --- Show save button for existing improved prompt ---
    if (st.session_state.get('improved_prompt') and 
        st.session_state['db_manager'] and 
        st.session_state.feedback_log):  # Only show if there's been feedback
        
        st.markdown("#### 💡 Previous Improved Prompt Available")
        st.info("You have a previously generated improved prompt that hasn't been saved to the database yet.")
        
        col_existing1, col_existing2, col_existing3 = st.columns([2, 1, 2])
        with col_existing2:
            save_existing_prompt = st.button(
                "💾 Save Previous Prompt", 
                key="save_existing_prompt_button",
                help="Save the previously generated improved prompt to database",
                type="secondary"
            )
        
        if save_existing_prompt and st.session_state.feedback_log:
            # Get the last feedback entry for context
            last_feedback = st.session_state.feedback_log[-1].get('feedback', 'No feedback text available')
            
            try:
                with st.spinner("💾 Saving previous improved prompt to database..."):
                    feedback_summary = last_feedback[:100]  # First 100 chars for title
                    success = save_improved_prompt(
                        prompt_text=st.session_state['improved_prompt'],
                        feedback_summary=feedback_summary,
                        use_case=st.session_state.get('use_case', 'Form 926'),
                        effectiveness_score=None,
                        feedback_requested=last_feedback
                    )
                    if success:
                        st.success("✅ Previous improved prompt saved to database and set as active!")
                        # Update current prompt ID for tracking
                        updated_prompt_data = get_latest_prompt(st.session_state.get('use_case', 'Form 926'))
                        if updated_prompt_data:
                            st.session_state['current_prompt_id'] = updated_prompt_data['PromptID']
                    else:
                        st.error("❌ Failed to save prompt to database")
            except Exception as e:
                st.error(f"❌ Database save error: {str(e)}")
        
        st.markdown("---")
    
    feedback = st.text_area("Suggest correction/feedback regarding the extracted data:")
    
    submit_feedback = st.button("Submit Feedback", key="feedback_button")
    
    if submit_feedback and feedback.strip():
        with st.spinner("🔄 Generating improved prompt..."):
            improved_prompt = asyncio.run(
                call_improvement_agent(
                    st.session_state['last_extraction'],
                    feedback,
                    st.session_state['last_prompt']
                )
            )
            st.session_state['improved_prompt'] = improved_prompt
            # Store the feedback that generated this prompt for database saving
            st.session_state['current_feedback'] = feedback
            # Reset the saved flag since this is a new prompt
            st.session_state['prompt_saved_to_db'] = False
        
        st.success("✅ Improved prompt generated!")
        
        # --- Run the extraction agent AGAIN with improved prompt ---
        with st.spinner("🔄 Re-running extraction with improved prompt..."):
            # Combine the improved prompt template with current document data
            complete_improved_prompt = f"{improved_prompt}\n\n" + (
                f"Current Task:\n"
                f"Extract the following columns from the PDF text:\n"
                f"Columns: {', '.join(columns)}\n"
                f"Instructions: {'; '.join(instructions)}\n"
                f"PDF Content:\n{pdf_text}\n"
            )
            
            improved_extraction = asyncio.run(
                call_extraction_agent(complete_improved_prompt, columns, instructions)
            )
            st.session_state['improved_extraction'] = improved_extraction
            
            # --- Save Improved Extraction as New Record in Database ---
            if st.session_state['db_manager']:
                try:
                    feedback_count = len(st.session_state.feedback_log) + 1
                    next_version = len(st.session_state.get('document_versions', [])) + 1
                    
                    with st.spinner("💾 Creating new version record for improved extraction..."):
                        # Create new document record for this iteration
                        new_document_id = st.session_state['db_manager'].insert_document_request(
                            file_name=pdf_file.name,
                            user_id=st.session_state.get('user_id', 'streamlit_user'),
                            source_type=f"Improvement_V{next_version}"
                        )
                        
                        if new_document_id:
                            # Update the new record with extraction results
                            success = st.session_state['db_manager'].update_document_master_by_id(
                                document_id=new_document_id,
                                extraction_status="Improved",
                                extraction_output=improved_extraction,
                                prompt_id=st.session_state.get('current_prompt_id'),
                                retry_count=feedback_count,
                                error_message=None,
                                comments=f"Version {next_version} - Iteration #{feedback_count} - Feedback: {feedback[:100]}{'...' if len(feedback) > 100 else ''}"
                            )
                            
                            if success:
                                # Track this new version
                                version_info = {
                                    'version': next_version,
                                    'document_id': new_document_id,
                                    'extraction_type': 'Improved',
                                    'feedback': feedback[:200]
                                }
                                
                                if 'document_versions' not in st.session_state:
                                    st.session_state['document_versions'] = []
                                st.session_state['document_versions'].append(version_info)
                                
                                # Update current document_id to point to latest version
                                st.session_state['current_document_id'] = new_document_id
                                
                                st.success(f"✅ New version (V{next_version}) created successfully (Document ID: {new_document_id})")
                                
                                # Show what was saved for verification
                                with st.expander("📋 New Version Details"):
                                    st.write(f"**New Document ID:** {new_document_id}")
                                    st.write(f"**Version:** {next_version}")
                                    st.write(f"**Status:** Improved")
                                    st.write(f"**Iteration:** {feedback_count}")
                                    st.write(f"**Prompt ID:** {st.session_state.get('current_prompt_id', 'N/A')}")
                                    st.write(f"**Feedback:** {feedback[:200]}{'...' if len(feedback) > 200 else ''}")
                                    st.write(f"**Extraction Output:** {len(improved_extraction)} characters")
                                    
                                    # Show all versions
                                    st.write("**All Versions:**")
                                    for v in st.session_state['document_versions']:
                                        st.write(f"- V{v['version']}: Document ID {v['document_id']} ({v['extraction_type']})")
                            else:
                                st.error("❌ Failed to update new document record with extraction results")
                        else:
                            st.error("❌ Failed to create new document record for improved extraction")
                except Exception as e:
                    st.error(f"❌ Database update failed for improved extraction: {str(e)}")
                    # Try to log the error in the database
                    try:
                        st.session_state['db_manager'].update_document_master_by_id(
                            document_id=st.session_state['document_id'],
                            extraction_status="Error",
                            error_message=f"Improved extraction error: {str(e)}",
                            retry_count=len(st.session_state.feedback_log) + 1,
                            comments=f"Error during improved extraction iteration {len(st.session_state.feedback_log) + 1}"
                        )
                    except:
                        pass

        # --- Log everything ---
        log_entry = {
            "old_extraction": st.session_state['last_extraction'],
            "feedback": feedback,
            "improved_prompt": improved_prompt,
            "improved_extraction": improved_extraction,
            "timestamp": datetime.now().isoformat(),
            "db_available": st.session_state['db_manager'] is not None,
            "document_id": st.session_state.get('document_id')
        }
        
        if st.session_state.get('current_prompt_id'):
            log_entry["prompt_id"] = st.session_state['current_prompt_id']
            
        st.session_state.feedback_log.append(log_entry)
        
        st.success("🎯 Extraction completed with improved prompt!")
    
    # --- Show improved prompt if it exists (persistent display) ---
    if st.session_state.get('improved_prompt'):
        st.markdown("#### Improved Extraction Prompt:")
        pro = json.loads(st.session_state['improved_prompt'] if isinstance(st.session_state['improved_prompt'], str) else '"{}"')
        # print("Type: ", type(st.session_state['improved_prompt']))
        st.code(pro['Prompt'], language="markdown")
        # st.code(st.session_state['improved_prompt'], language="markdown")
        
        # --- Show save status ---
        if st.session_state.get('prompt_saved_to_db'):
            st.success("✅ This prompt has been saved to the database!")
        
        # --- Database Save Button (only show if database is available and not already saved) ---
        if st.session_state['db_manager'] and not st.session_state.get('prompt_saved_to_db'):
            st.markdown("---")
            col_save1, col_save2, col_save3 = st.columns([2, 1, 2])
            
            with col_save2:
                save_prompt_to_db = st.button(
                    "💾 Save Prompt to Database", 
                    key="save_prompt_db_button",
                    help="Save this improved prompt to database and set it as active for future use",
                    type="primary"
                )
            
            if save_prompt_to_db:
                # Get the feedback that was used to generate this prompt
                feedback_for_prompt = st.session_state.get('current_feedback', 'No feedback available')
                
                try:
                    with st.spinner("💾 Saving improved prompt to database..."):
                        feedback_summary = feedback_for_prompt[:100]  # First 100 chars for title
                        
                        # Show what we're attempting to save
                        with st.expander("📝 Save Details", expanded=False):
                            st.write(f"**Use Case:** {st.session_state.get('use_case', 'Form 926')}")
                            st.write(f"**Feedback Summary:** {feedback_summary}")
                            st.write(f"**Prompt Length:** {len(st.session_state['improved_prompt'])} characters")
                            st.write(f"**Feedback Length:** {len(feedback_for_prompt)} characters")
                        
                        success = save_improved_prompt(
                            prompt_text=st.session_state['improved_prompt'],
                            feedback_summary=feedback_summary,
                            use_case=st.session_state.get('use_case', 'Form 926'),
                            effectiveness_score=None,  # Could be set based on user rating
                            feedback_requested=feedback_for_prompt  # Store the full feedback text
                        )
                        
                        if success:
                            st.success("✅ Improved prompt saved to database and set as active!")
                            # Update current prompt ID for tracking
                            updated_prompt_data = get_latest_prompt(st.session_state.get('use_case', 'Form 926'))
                            if updated_prompt_data:
                                st.session_state['current_prompt_id'] = updated_prompt_data['PromptID']
                                st.info(f"🔄 Active prompt updated to ID: {updated_prompt_data['PromptID']}")
                            
                            # Mark that this prompt has been saved to avoid duplicate saves
                            st.session_state['prompt_saved_to_db'] = True
                            
                            # Force a rerun to update the UI
                            st.rerun()
                        else:
                            st.error("❌ Failed to save prompt to database - stored procedure returned False")
                            
                except Exception as e:
                    st.error(f"❌ Database save error: {str(e)}")
                    # Show detailed error information for debugging
                    with st.expander("🔍 Error Details"):
                        st.code(str(e))
                        st.write("**Attempted to save:**")
                        st.write(f"- Use Case: {st.session_state.get('use_case', 'Form 926')}")
                        st.write(f"- Feedback Summary: {feedback_for_prompt[:100]}")
                        st.write(f"- Prompt Length: {len(st.session_state['improved_prompt'])} characters")
                        
                        # Test database connection
                        if st.session_state['db_manager']:
                            conn_test = st.session_state['db_manager'].test_connection()
                            st.write(f"- Database Connection: {'✅ OK' if conn_test else '❌ Failed'}")
        else:
            st.info("💡 Database not available - prompt generated but not saved")
    
    elif submit_feedback and not feedback.strip():
        st.warning("⚠️ Please provide feedback before submitting.")

    # --- Show side-by-side results (if improved available) ---
    if st.session_state['improved_extraction']:
        st.markdown("### Side-by-Side Extraction Results")
        col1, col2 = st.columns(2)
        with col1:
            st.markdown("**Original Extraction**")
            original_df = json_to_dataframe(st.session_state['last_extraction'], columns)
            st.dataframe(original_df, use_container_width=True)
            
            with st.expander("🔍 View Raw JSON"):
                st.code(st.session_state['last_extraction'], language="json")
                
        with col2:
            st.markdown("**Extraction With Improved Prompt**")
            improved_df = json_to_dataframe(st.session_state['improved_extraction'], columns)
            st.dataframe(improved_df, use_container_width=True)
            
            with st.expander("🔍 View Raw JSON"):
                st.code(st.session_state['improved_extraction'], language="json")

    # --- Excel Export Section (Always Available) ---
    st.markdown("### 📥 Export to Excel")
    
    if st.session_state.get('last_extraction'):
        # Get the current best extraction (improved if available, otherwise original)
        current_extraction = st.session_state.get('improved_extraction') or st.session_state.get('last_extraction')
        
        # Show preview of current data
        try:
            if isinstance(current_extraction, str):
                preview_data = json.loads(current_extraction)
            else:
                preview_data = current_extraction
            
            preview_df = pd.DataFrame([preview_data])
            # Ensure all columns are present
            for col in columns:
                if col not in preview_df.columns:
                    preview_df[col] = ""
            preview_df = preview_df[columns]
            
            st.markdown("**Current Extraction Preview:**")
            st.dataframe(preview_df, use_container_width=True)
            
        except (json.JSONDecodeError, Exception) as e:
            st.warning(f"Preview not available: {str(e)}")
        
        # Download buttons in columns
        col1, col2, col3 = st.columns(3)
        
        with col1:
            # Always available: Download original extraction
            excel_data_original = json_to_excel(st.session_state['last_extraction'], columns)
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            
            st.download_button(
                label="📋 Original Extraction",
                data=excel_data_original,
                file_name=f"original_extraction_{timestamp}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                key="download_original_main",
                help="Download the first extraction result"
            )
        
        with col2:
            # Available if improved extraction exists
            if st.session_state.get('improved_extraction'):
                excel_data_improved = json_to_excel(st.session_state['improved_extraction'], columns)
                
                st.download_button(
                    label="🔄 Latest Improved",
                    data=excel_data_improved,
                    file_name=f"improved_extraction_{timestamp}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    key="download_improved_main",
                    help="Download the most recent improved extraction"
                )
            else:
                st.info("💡 Submit feedback to get improved version")
        
        with col3:
            # Only show "Best Available" if no improved extraction exists
            # (otherwise it would be identical to "Latest Improved")
            if not st.session_state.get('improved_extraction'):
                excel_data_current = json_to_excel(current_extraction, columns)
                
                st.download_button(
                    label="⚡ Best Available",
                    data=excel_data_current,
                    file_name=f"best_extraction_{timestamp}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    key="download_best_main",
                    help="Download the original extraction (no improvements yet)"
                )
            else:
                st.info("💡 Use 'Latest Improved' for best results")
        
        # Show feedback history downloads if available
        if st.session_state.feedback_log:
            st.markdown("---")
            st.markdown("**📚 Download Previous Iterations**")
            
            # Show last few feedback iterations
            recent_feedback = st.session_state.feedback_log[-3:]  # Last 3 iterations
            
            for i, feedback_entry in enumerate(reversed(recent_feedback)):
                iteration = len(st.session_state.feedback_log) - i
                
                col_iter1, col_iter2 = st.columns([3, 1])
                
                with col_iter1:
                    st.write(f"**Iteration {iteration}:** {feedback_entry.get('feedback', 'N/A')[:100]}...")
                
                with col_iter2:
                    if 'improved_extraction' in feedback_entry:
                        excel_iteration = json_to_excel(feedback_entry['improved_extraction'], columns)
                        
                        st.download_button(
                            label=f"📥 V{iteration}",
                            data=excel_iteration,
                            file_name=f"extraction_v{iteration}_{timestamp}.xlsx",
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                            key=f"download_iteration_{iteration}",
                            help=f"Download extraction from iteration {iteration}"
                        )
    else:
        st.info("🔍 No extraction data available. Please upload files and run extraction first.")
    
    # Additional export options
    if st.session_state.feedback_log:
        st.markdown("---")
        with st.expander("🔽 Advanced Export Options"):
            st.markdown("**Export All Iterations as Single File**")
            
            if st.button("📊 Create Multi-Sheet Excel", key="multi_sheet_export"):
                # Create Excel with multiple sheets
                output = io.BytesIO()
                with pd.ExcelWriter(output, engine='openpyxl') as writer:
                    
                    # Original extraction
                    if st.session_state.get('last_extraction'):
                        try:
                            original_data = json.loads(st.session_state['last_extraction']) if isinstance(st.session_state['last_extraction'], str) else st.session_state['last_extraction']
                            original_df = pd.DataFrame([original_data])
                            for col in columns:
                                if col not in original_df.columns:
                                    original_df[col] = ""
                            original_df = original_df[columns]
                            original_df.to_excel(writer, sheet_name='Original', index=False)
                        except:
                            pass
                    
                    # Each feedback iteration
                    for i, feedback_entry in enumerate(st.session_state.feedback_log):
                        if 'improved_extraction' in feedback_entry:
                            try:
                                iter_data = json.loads(feedback_entry['improved_extraction']) if isinstance(feedback_entry['improved_extraction'], str) else feedback_entry['improved_extraction']
                                iter_df = pd.DataFrame([iter_data])
                                for col in columns:
                                    if col not in iter_df.columns:
                                        iter_df[col] = ""
                                iter_df = iter_df[columns]
                                iter_df.to_excel(writer, sheet_name=f'Iteration_{i+1}', index=False)
                            except:
                                continue
                
                timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                st.download_button(
                    label="⬇️ Download Multi-Sheet Excel",
                    data=output.getvalue(),
                    file_name=f"all_extractions_{timestamp}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    key="download_multi_sheet"
                )

if st.sidebar.checkbox("Show Feedback Log", value=True):
    st.sidebar.markdown("### Feedback Log")
    st.sidebar.write(st.session_state.feedback_log)

# --- Database Status Sidebar ---
st.sidebar.markdown("---")
st.sidebar.markdown("### Database Status")
if st.session_state.get('db_manager'):
    st.sidebar.success("✅ Connected")
    
    # Show document status and version information
    if st.session_state.get('document_id') or st.session_state.get('document_versions'):
        st.sidebar.markdown("### Document Versions")
        
        # Show version history
        if st.session_state.get('document_versions'):
            st.sidebar.info(f"📊 Total Versions: {len(st.session_state['document_versions'])}")
            
            if files_uploaded:
                st.sidebar.caption(f"File: {pdf_file.name}")
            
            # Show each version
            for version in st.session_state['document_versions']:
                version_num = version['version']
                doc_id = version['document_id']
                extraction_type = version['extraction_type']
                
                if version_num == 1:
                    st.sidebar.success(f"📄 V{version_num}: ID {doc_id} ({extraction_type})")
                else:
                    st.sidebar.info(f"🔄 V{version_num}: ID {doc_id} ({extraction_type})")
                    if version.get('feedback'):
                        st.sidebar.caption(f"Feedback: {version['feedback'][:50]}...")
            
            # Show current active version
            if st.session_state.get('current_document_id'):
                st.sidebar.markdown("**Active Version:**")
                st.sidebar.warning(f"🎯 Current: Document ID {st.session_state['current_document_id']}")
        
        elif st.session_state.get('document_id'):
            # Fallback for single document (no versioning yet)
            st.sidebar.info(f"📄 Document ID: {st.session_state['document_id']}")
            if files_uploaded:
                st.sidebar.caption(f"File: {pdf_file.name}")
                feedback_count = len(st.session_state.feedback_log)
                if feedback_count > 0:
                    st.sidebar.caption(f"Iterations: {feedback_count}")
    
    if st.session_state.get('current_prompt_id'):
        st.sidebar.info(f"📋 Prompt ID: {st.session_state['current_prompt_id']}")
    
    # Show selected use case
    if st.session_state.get('use_case'):
        st.sidebar.info(f"📝 Use Case: {st.session_state['use_case']}")
    
    # Button to refresh prompt from database
    if st.sidebar.button("🔄 Refresh Prompt", help="Get latest prompt from database"):
        try:
            db_prompt_data = get_latest_prompt(st.session_state.get('use_case', 'Form 926'))
            if db_prompt_data:
                st.session_state['current_prompt_id'] = db_prompt_data['PromptID']
                st.sidebar.success(f"✅ Refreshed to Prompt ID: {db_prompt_data['PromptID']}")
                st.rerun()
            else:
                st.sidebar.warning("⚠️ No active prompts found")
        except Exception as e:
            st.sidebar.error(f"❌ Refresh failed: {str(e)}")
else:
    st.sidebar.warning("⚠️ Database Offline")
    st.sidebar.caption("Using local prompts only")
    
    # Show connection attempt info for debugging
    if st.sidebar.button("🔧 Test Connection", help="Try to connect to database"):
        try:
            test_db = DatabaseManager()
            if test_db.test_connection():
                st.sidebar.success("✅ Connection successful! Refresh page.")
            else:
                st.sidebar.error("❌ Connection failed")
                with st.sidebar.expander("🔍 Debug Info"):
                    st.caption(test_db.get_connection_info())
        except Exception as e:
            st.sidebar.error(f"❌ Connection error: {str(e)}")

if not files_uploaded:
    st.info("Please upload both files and confirm to start the extraction/feedback process.")
